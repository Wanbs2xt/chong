import time
import requests
from bs4 import BeautifulSoup, NavigableString
from playwright.sync_api import Playwright, sync_playwright, expect
import pandas as pd
import urllib.parse
import re
from bpftUI import BaiduPanFilesTransfers
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


def set_excel_format(path):
    wb = load_workbook(path)
    ws = wb.active

    # 设置行高，跳过首行
    for i, row in enumerate(ws.iter_rows(), start=1):
        if i > 1:  # 跳过首行
            ws.row_dimensions[row[0].row].height = 187.2

    # 设置列宽
    ws.column_dimensions[get_column_letter(2)].width = 22.33  # 名称列宽，对应"B"列
    ws.column_dimensions[get_column_letter(3)].width = 34.33  # 豆瓣名称列宽，对应"C"列
    ws.column_dimensions[get_column_letter(5)].width = 82.56  # 豆瓣信息列宽，对应"E"列
    ws.column_dimensions[get_column_letter(6)].width = 82.56  # 百度网盘分享外链列宽，对应"F"列
    ws.column_dimensions[get_column_letter(9)].width = 82.56  # 百度链接列宽，对应"I"列

    # 设置EFI列自动换行
    alignment = Alignment(wrap_text=True)
    for col in ['E', 'F', 'I']:
        for cell in ws[col]:
            cell.alignment = alignment

    wb.save(path)


def link_test(url):
    session = requests.Session()
    head = {
        'User-agnet': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
    }
    res = session.get(url, headers=head)
    soup = BeautifulSoup(res.text, 'html.parser')
    text = str(soup.title)
    return text


def save_to_baidu(classtype, total_title):
    app = BaiduPanFilesTransfers()
    app.entry_cookie.insert(0, 'cookie')
    if classtype == '电影':
        app.entry_folder_name.insert(0, total_title)
    else:
        app.create_dir(total_title)
        app.entry_folder_name.insert(0, total_title)

    app.text_links.insert(1.0, "你的链接")


def search_movie(page, filename):
    url1 = 'http://m.kkkob.com/apps/index.html?id=211229kl'
    url2 = 'http://ysxjjkl.souyisou.top/'
    page.goto(url1)
    page.query_selector('.input-wap #search').fill(filename)
    page.query_selector('#submitSearch').click()
    page.wait_for_load_state('networkidle')
    pageinfo1 = page.content()
    page.goto(url2)
    page.query_selector('.input-wap #search').fill(filename)
    page.query_selector('#submitSearch').click()
    page.wait_for_load_state('networkidle')
    pageinfo2 = page.content()
    soup = BeautifulSoup(pageinfo1, 'html.parser')
    soup2 = BeautifulSoup(pageinfo2, 'html.parser')
    baidu_link = [
        (a['href'].split('?')[0],
         re.search(r'提取码：(\w+)', next_sibling.get_text()).group(1) if next_sibling and re.search(r'提取码：(\w+)',
                                                                                                    next_sibling.get_text()) else None)
        for a in soup.select('.access-box .info a') if
        'pan.baidu.com' in a['href'] and (next_sibling := a.find_next_sibling('div'))
    ]

    baidu_link2 = [
        (a['href'].split('?')[0],
         re.search(r'提取码：(\w+)', next_sibling.get_text()).group(1) if next_sibling and re.search(r'提取码：(\w+)',
                                                                                                    next_sibling.get_text()) else None)
        for a in soup2.select('.access-box .info a') if
        'pan.baidu.com' in a['href'] and (next_sibling := a.find_next_sibling('div'))
    ]
    combined_links = list(set(baidu_link + baidu_link2))
    datastr = ''  # 初始化datastr
    if combined_links:
        for link, password in combined_links:
            try:
                res = link_test(link)
                if '提取码' in res or '下载' in res:
                    datastr += link + (f'?pwd={password}' if password else '') + '\n'  # 将符合条件的链接加入到datastr中
            except:
                continue
    return datastr.strip() if datastr else None


def read_excel_file(path):
    df = pd.read_excel(path)
    return df


def get_page_content(page, url):
    page.goto(url)
    soup = BeautifulSoup(page.content(), 'html.parser')
    return soup


def get_movie_type(soup):
    class_type = None
    elements = soup.select('.result .title span')
    if elements:
        element = elements[0].text
        if element:
            class_type = element.strip('[]')
            if class_type not in ('电影', '电视剧'):
                class_type = '电视剧'
    return class_type


def get_movie_info_link(soup):
    href_links = [a['href'] for a in soup.select('.sc-bZQynM a')]
    return href_links[0] if href_links else None


def get_movie_title_and_year(soup, file_name):
    title = soup.find('span', property="v:itemreviewed")
    title_text = title.text
    start = file_name[0]
    end = file_name[-1]
    totalname = title_text[title_text.find(start):title_text.find(end) + 1]
    finall_name = re.sub(r'[^\u4e00-\u9fff\w]+', '', totalname)
    if file_name in finall_name:
        try:
            year = soup.find('span', class_='year').text
        except:
            year = ''
        total_title = totalname + year
        return total_title, year
    return None, None


def get_movie_info(soup):
    info_div = soup.find('div', {'id': 'info'})
    movie_info = ''

    # 遍历info_div的所有直接子元素
    for child in info_div:

        # 检查当前子元素是否是NavigableString（没有标签的字符串）
        if isinstance(child, NavigableString):
            text = child.strip()
            # 如果文本不为空，添加到结果字符串
            # 如果文本以冒号结尾（说明是属性名），在冒号后添加一个空格
            if text:
                if text.endswith(':'):
                    text += ' '
                movie_info += text
                if '更多...' in movie_info:
                    movie_info = movie_info.replace('更多...', '')

        # 检查当前子元素是否是HTML标签
        elif child.name:

            # 如果是br标签，添加换行
            if child.name == 'br':
                movie_info += '\n'
            # 如果是span标签，添加标签的文本内容
            elif child.name == 'span':
                text = child.text.strip()
                if text:
                    movie_info += text
                    if '更多...' in movie_info:
                        movie_info = movie_info.replace('更多...', '')
            # 如果是a标签，添加此标签的文本内容和链接
            elif child.name == 'a':
                link_text = child.text.strip()
                movie_info += f"{link_text}"
    return movie_info


def update_excel_file(df, file_name, total_title, classtype, path, info_text, baidulink):
    row_index = df[df['名称'] == file_name].index[0]
    df.loc[row_index, '豆瓣名称'] = total_title
    df.loc[row_index, '类型'] = classtype
    df.loc[row_index, '豆瓣信息'] = info_text
    df.loc[row_index, '百度链接'] = baidulink
    df.to_excel(path, index=False)


def run(playwright: Playwright) -> None:
    df = read_excel_file(path)
    with playwright.chromium.launch(headless=False) as browser:
        with browser.new_context() as context:
            page = context.new_page()
            context.set_default_navigation_timeout(60000)  # 设置默认的导航超时时间为60秒
            context.set_default_timeout(60000)
            b_column = df['名称']
            for file_name in b_column:
                baidulink = search_movie(page, file_name)
                if not baidulink:
                    continue
                encoded_string = urllib.parse.quote(file_name)
                soup = get_page_content(page, f'https://www.douban.com/search?q={encoded_string}')
                classtype = get_movie_type(soup)
                soup = get_page_content(page,
                                        f'https://search.douban.com/movie/subject_search?search_text={encoded_string}&cat=1002')
                movie_link = get_movie_info_link(soup)
                if movie_link:
                    soup = get_page_content(page, movie_link)
                    total_title, year = get_movie_title_and_year(soup, file_name)
                    info_text = get_movie_info(soup)
                    if total_title and year:
                        update_excel_file(df, file_name, total_title, classtype,
                                          path, info_text, baidulink)
                        set_excel_format(path)
            set_excel_format(path)


with sync_playwright() as playwright:
    path = r"文件路径"
    run(playwright)
